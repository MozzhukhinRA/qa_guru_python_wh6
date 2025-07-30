import pandas as csv
import os.path
from io import BytesIO
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from paths import TMP_DIR


def test_archiving_files():
    files_for_zip = [f'{TMP_DIR}/example.csv', f'{TMP_DIR}/example.pdf', f'{TMP_DIR}/example.xlsx']
    with ZipFile(f'{TMP_DIR}/test_archive.zip', 'w') as zipp:
        for zipf in files_for_zip:
            zipp.write(zipf, arcname=os.path.basename(zipf))


def test_search_file_into_archive():
    with ZipFile(f'{TMP_DIR}/test_archive.zip') as zipp:
        assert 'example.csv' in zipp.namelist()
        assert 'example.pdf' in zipp.namelist()
        assert 'example.xlsx' in zipp.namelist()
        assert os.path.getmtime(f'{TMP_DIR}/test_archive.zip') == 1753867417.9110734


def test_read_file_pdf_into_archive():
    with ZipFile(f'{TMP_DIR}/test_archive.zip', 'r') as zipp:
        unit = PdfReader(BytesIO(zipp.read('example.pdf')))
        assert "Document file type: PDF" in unit.get_page(0).extract_text()
        assert len(unit.pages) == 1
        assert "New Zealand." in unit.pages[0].extract_text()


def test_read_file_xlsx_into_archive():
    with ZipFile(f'{TMP_DIR}/test_archive.zip', 'r') as zipp:
        unit = load_workbook(BytesIO(zipp.read('example.xlsx'))).active
        assert "Month" in unit.cell(row=12, column=1).value


def test_read_file_csv_into_archive():
    with ZipFile(f'{TMP_DIR}/test_archive.zip', 'r') as zipp:
        unit = csv.read_csv(BytesIO(zipp.read('example.csv')))
        assert "123" in list(unit.columns)
